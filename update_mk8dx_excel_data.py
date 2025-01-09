"""
This script processes images of Mario Kart 8 Deluxe final times
then extracts the lap times and other data.

Author: Stephen Haptonstahl
Email: srh@haptonstahl.org
Source: https://github.com/shaptonstahl/mk8dx-lap-time-tracker
"""

import tempfile
from PIL import Image, ImageOps  # noqa: F401
from pathlib import Path
import os
import datetime as dt
import subprocess
from mindee import Client, AsyncPredictResponse, product
from dotenv import load_dotenv
import openpyxl
import openpyxl.worksheet
import openpyxl.worksheet.table
from openpyxl.utils.dataframe import dataframe_to_rows
import polars as pl
import pickle
import argparse
from pprint import pprint
import re
import openai
from pydantic import BaseModel
import base64

# Make Mindee and OpenAI API keys accessible
load_dotenv()

# Set some OpenAI variables
openai_api_key = os.getenv("OPENAI_API_KEY")
if openai_api_key is None:
    raise ValueError("OpenAI API key is not set in environment variables.")

openai_url = "https://api.openai.com/v1/chat/completions"
openai_headers = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {openai_api_key}",
}


# Image kind identification uses ocrs without processing other than cropping.
image_kind_identification_parameters = {
    "order_to_check": [1, 3, 2],
    1: {
        "title": "lower_right_OK",
        "box": [1170, 646, 1225, 682],
        "expected_value": ' == "OK"',
    },
    2: {
        "title": "last_digit_top_lap_2",
        "box": [1162, 220, 1183, 253],
        "expected_value": " != None",
    },
    3: {
        "title": "lower_right_lap_2_number",
        "box": [1020, 456, 1044, 480],
        "expected_value": ' == "2"',
    },
}


class kind_1_image(BaseModel):
    """
    Defines the return type for a call to the OpenAI API
    """

    racer: str
    character_name: str
    vehicle: str
    wheels: str
    glider: str
    track: str
    overall_time: str
    lap_1_time: str
    lap_2_time: str
    lap_3_time: str
    nationality: str


class kind_2_image(BaseModel):
    """
    Defines the return type for a call to the OpenAI API
    """

    character_name: str
    vehicle: str
    wheels: str
    glider: str
    overall_time: str
    lap_1_time: str
    lap_2_time: str
    lap_3_time: str
    nationality: str


class kind_3_image(BaseModel):
    """
    Defines the return type for a call to the OpenAI API
    """

    racer: str
    character_name: str
    vehicle: str
    wheels: str
    glider: str
    overall_time: str
    lap_1_time: str
    lap_2_time: str
    lap_3_time: str
    nationality: str
    racer_2: str
    character_name_2: str
    vehicle_2: str
    wheels_2: str
    glider_2: str
    overall_time_2: str
    lap_1_time_2: str
    lap_2_time_2: str
    lap_3_time_2: str
    nationality_2: str


# schema: prompts[llm_platform][image_kind] = 'prompt'
prompts = {
    "openai": {
        1: """
You are a Mario Kart 8 Deluxe player. Look at the included image. 
Identify the nationality of the player by looking at the national
flag under the racer's name. Identify the Mario Kart character by
looking at the character's head to the left of the racer's name.
Pay attention to whether the character is a baby or not because
that may change its name.
Respond in json with the following fields:
""",
        2: """
You are a Mario Kart 8 Deluxe player. Look at the included image. 
Identify the nationality of the player by looking at the national
flag to the right of the racer's name and to the left of the 
overall_time. Identify the Mario Kart character by looking at the 
character's head to the left of the racer's name.
Pay attention to whether the character is a baby or not because
that may change its name.
Respond in json with the following fields:
""",
        3: """
You are a Mario Kart 8 Deluxe player. Look at the included image.
There are two blocks of text and numbers, a top block and a bottom 
block. The top block lists information about player one and the bottom
block lists information about player two.

For each player, idenitfy:

-  The nationality of the player by looking at the national flag to
the right of the players's name and to the left of the overall_time.
Field: "nationality" or "nationality_2". If the flag is the United
States flag, set the field to 'USA'.

- The Mario Kart character by looking at the character's head to the
left of the player's name. Pay attention to whether the character is 
a baby or not because that may change its name.
Field: "character_name" or "character_name_2"

- The name of the player. Field: "racer" or "racer_2"

- Overall time. The overall time is to the right of the nationality 
flag for each player.

- Times for lap 1, lap 2, and lap 3

- Vehicle

- Wheels

- Glider
""",
    }
}

prompts["openai"][1] = prompts["openai"][1] + ", ".join(
    [k for k, v in kind_1_image.model_fields.items()]
)
prompts["openai"][2] = prompts["openai"][2] + ", ".join(
    [k for k, v in kind_2_image.model_fields.items()]
)


def encode_image_for_api_call(image_path: Path) -> str:
    """
    Given the location of an image, encode as a string suitable for inclusion
    in an API call, particularly to OpenAI.
    """
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode("utf-8")


# This defines how to extract fields from the image once the kind has been identified.
# Each kind has a list of extractions that is applied in order. Each extraction is a
# dictionary.
image_processing_parmeters = {
    1: [
        {
            "framework": "openai",
            "box": [1, 1, 1280, 720],
            "transformations": [],
            "fields": [k for k, v in kind_1_image.model_fields.items()],
        }
    ],
    2: [
        {
            "framework": "openai",
            "box": [1, 1, 1280, 720],
            "transformations": [],
            "fields": [k for k, v in kind_2_image.model_fields.items()],
        }
    ],
    3: [
        {
            "framework": "openai",
            "box": [1, 1, 1280, 720],
            "transformations": [],
            "fields": [k for k, v in kind_3_image.model_fields.items()],
        }
    ],
    11: [
        {
            "framework": "mindee",
            "box": [1, 1, 1280, 720],
            "transformations": [],
            "endpoint_name": "mk8dx_screen_capture_kind_1",
            "endpoint_version": "1",
            "fields": [
                "racer",
                "track",
                "vehicle",
                "wheels",
                "glider",
                "overall_time",
                "lap_1_time",
                "lap_2_time",
                "lap_3_time",
            ],
        }
    ],
    21: [
        {
            "framework": "mindee",
            "box": [1, 1, 1280, 720],
            "transformations": [],
            "endpoint_name": "mk8dx_screen_capture_kind_2",
            "endpoint_version": "1",
            "fields": ["character_name", "vehicle", "wheels", "glider"],
        },
        {
            "framework": "ocrs",
            "box": [1047, 145, 1200, 210],
            "transformations": [
                {"transformation": "grayscale"},
                {
                    "transformation": "autocontrast",
                    "parameters": {"cutoff": 0.5, "preserve_tone": True},
                },
            ],
            "fields": ["lap_1_time"],
        },
        {
            "framework": "ocrs",
            "box": [1047, 222, 1200, 253],
            "transformations": [
                {"transformation": "grayscale"},
                {
                    "transformation": "autocontrast",
                    "parameters": {"cutoff": 0.5, "preserve_tone": True},
                },
            ],
            "fields": ["lap_2_time"],
        },
        {
            "framework": "ocrs",
            "box": [1047, 264, 1200, 298],
            "transformations": [
                {"transformation": "grayscale"},
                {
                    "transformation": "autocontrast",
                    "parameters": {"cutoff": 0.5, "preserve_tone": True},
                },
            ],
            "fields": ["lap_3_time"],
        },
        {
            "framework": "ocrs",
            "box": [1039, 113, 1210, 152],
            "transformations": [
                {"transformation": "grayscale"},
                {"transformation": "invert"},
                {
                    "transformation": "autocontrast",
                    "parameters": {"cutoff": 0.75, "preserve_tone": True},
                },
            ],
            "fields": ["overall_time"],
        },
    ],
    31: [
        {
            "framework": "mindee",
            "box": [1, 1, 1280, 720],
            "transformations": [],
            "endpoint_name": "mk8dx_screen_capture_kind_3",
            "endpoint_version": "1",
            "fields": [
                "racer",
                "vehicle",
                "wheels",
                "glider",
                "overall_time",
                "lap_1_time",
                "lap_2_time",
                "lap_3_time",
                "opponent",
                "vehicle",
                "wheels",
                "glider",
                "overall_time",
                "lap_1_time",
                "lap_2_time",
                "lap_3_time",
            ],
        },
        {
            "framework": "ocrs",
            "box": [1, 1, 1280, 720],
            "transformations": [],
            "fields": ["something"],
        },
    ],
    41: [
        {"framework": "mindee", "box": [1, 1, 1280, 720]},
        {"framework": "ocrs", "fields": ["something"]},
    ],
}


def validate_extraction(extraction: dict) -> bool:
    """
    Given an extraction dictionary, return True if the extraction
    is correctly formatted and False otherwise.
    """
    if not isinstance(extraction, dict):
        return False
    if not all([x in extraction for x in ["framework", "fields"]]):
        return False
    if (
        "mindee" not in extraction["framework"]
        and "ocrs" not in extraction["framework"]
    ):
        return False
    if extraction["framework"] == "mindee":
        if not all([x in extraction for x in ["endpoint_name", "endpoint_version"]]):
            return False
    if not isinstance(extraction["endpoint_version"], str):
        return False
    return True


def validate_all_extractions(extractions: list[dict]) -> bool:
    """
    Given a list of extractions, return True if all of the extractions
    are correctly formatted and False otherwise.
    """
    return all([validate_extraction(x) for x in extractions])


def transform_image_and_save(
    image_path: Path, box: list[int] = None, transformations: list[dict] = None
) -> Path:
    """
    Given an image file and a list of transformations, perform the transformations
    and return the Path to a temporary file with the transformed image.

    transformations = a list of dictionaries, each a transformation from the ImageOps module.

    Examples:
    transformations =
    [
        {'transformation': 'autocontrast',
         'parameters': {'cutoff': 0.5, 'preserve_tone': True}},
        {'transformation': 'expand',
         'parameters': {border: 10}},
        {'transformation': 'grayscale'}} # Note that if there are no paremeters, then the dictionary can be simplified.
    ]

    List of transformations: autocontrast, colorize, scale, SuppoetGetMesh, deform, equalize,
    expand, flip, grayscale, invert, mirror, posterize, solarize, exif_transpose
    Refer to https://pillow.readthedocs.io/en/stable/reference/ImageOps.html for the complete list.

    Cropping with the specified box is always the first transformation.
    """
    with Image.open(image_path) as img:
        if box is not None:
            img = img.crop(box)
        if transformations is not None:
            for t in transformations:
                if "parameters" in t:
                    img = eval("ImageOps." + t["transformation"])(
                        img, **t["parameters"]
                    )
                else:
                    img = eval("ImageOps." + t["transformation"])(img)

        if debug:
            temp_dir = Path(os.getcwd())
        else:
            temp_dir = None

        with tempfile.NamedTemporaryFile(
            delete=False, suffix=".jpg", dir=temp_dir
        ) as temporary_file:
            temporary_file_name = temporary_file.name
            img.save(temporary_file_name)
    return Path(temporary_file_name)


def ocr_image(
    image_file: Path,
    framework: str,
    kind: int = None,
    endpoint_name: str = None,
    endpoint_version: str = None,
) -> dict:
    """
    Given an image Path (which may be cropped and transformed)
    and a framework, return the text ocr'ed from that box in a
    dictionary where key is a field name and value is a string.

    box = a list of integers representing pixels: [x1, y1, x2, y2]
    framework = a string: 'mindee' or 'ocrs'
    """

    if framework == "mindee":
        mindee_client = Client(api_key=os.getenv("MINDEE_API_KEY"))
        input_doc = mindee_client.source_from_path(image_file)
        endpoint = mindee_client.create_endpoint(
            account_name="polimath",
            endpoint_name=endpoint_name,
            version=endpoint_version,
        )
        result: AsyncPredictResponse = mindee_client.enqueue_and_parse(
            product.GeneratedV1, input_doc, endpoint=endpoint
        )
        del mindee_client
        return {
            k: v.value for k, v in result.document.inference.prediction.fields.items()
        }
    elif framework == "openai":
        openai_client = openai.OpenAI(api_key=openai_api_key)
        base64_image = encode_image_for_api_call(image_file)
        openai_response = openai_client.beta.chat.completions.parse(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompts["openai"][kind]},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/jpeg;base64,{base64_image}"
                            },
                        },
                    ],
                }
            ],
            max_tokens=300,
            response_format=eval(f"kind_{kind}_image"),
        )

        return dict(openai_response.choices[0].message.parsed)
    elif framework == "ocrs":
        ocr_result = subprocess.run(["ocrs", image_file], stdout=subprocess.PIPE)
        return {"ocr_result": ocr_result.stdout.decode("utf-8").strip()}
    else:
        return {}


def identify_kind(image_path: Path) -> int:
    """
    Given an image, return the kind of image.
    """
    if Image.open(image_path).size != (1280, 720):
        return 0
    for kind_to_check in image_kind_identification_parameters["order_to_check"]:
        parameters = image_kind_identification_parameters[kind_to_check]
        transformed_file_name = transform_image_and_save(
            image_path, box=parameters["box"]
        )
        ocr_result = ocr_image(transformed_file_name, framework="ocrs")["ocr_result"]  # noqa: F841
        condition_to_check = "ocr_result" + parameters["expected_value"]
        if debug:
            print(f"Checking condition: {condition_to_check}")
        if eval(condition_to_check):
            return kind_to_check
    return 0


def extract_fields_from_image(image_path: Path) -> dict:
    """
    Given an image file, return a dictionary of fields extracted from the image.
    """
    kind = identify_kind(image_path)
    if kind == 0:
        return {}

    output = {}
    # 'kind_extractions' is a list for the given kind.
    kind_extractions = image_processing_parmeters[kind]

    for parameters in kind_extractions:  # 'parameters' is a dict for each extraction.
        temp_file_name = transform_image_and_save(
            image_path,
            box=parameters.get("box", None),
            transformations=parameters.get("transformations", None),
        )
        ocr_output = ocr_image(
            temp_file_name,
            framework=parameters["framework"],
            kind=kind,
            endpoint_name=parameters.get("endpoint_name", None),
            endpoint_version=parameters.get("endpoint_version", None),
        )

        if parameters["framework"] == "ocrs":
            output.update({parameters["fields"][0]: ocr_output["ocr_result"]})
        elif parameters["framework"] in ["mindee", "openai"]:
            output.update(ocr_output)
        else:
            pass

        os.remove(temp_file_name)

    # overall_times = calc_overall_times(output)
    # breakpoint()
    # output["overall_time"] = overall_times[0]
    # if len(overall_times) == 2:
    #     output["overall_time_2"] = overall_times[1]
    #
    # ADD CHECK THAT OVERALL TIMES ARE THE SUM OF THE LAP TIMES
    #
    output["kind"] = kind
    output["image_file_name"] = image_path.name
    output["datetime_processed"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
    output["race_end_datetime"] = dt.datetime.strptime(
        image_path.stem[0:14], "%Y%m%d%H%M%S"
    ).strftime("%Y-%m-%d %H:%M:%S")

    return output


def read_list_of_files(folder: Path) -> set[Path]:
    """
    Given a folder Path object, return a list of file Path objects
    of the *.jpg files in that folder. Does not go into subfolders.
    """
    all_files_in_folder = Path(folder).glob("*.jpg")
    all_jpg_file_paths = [x for x in all_files_in_folder if x.is_file()]
    return set(all_jpg_file_paths)


def load_cache_and_list_of_paths_to_process(
    folder_to_process: Path, cache_file_path: Path
) -> tuple[pl.DataFrame, set[Path]]:
    """
    Given a folder to process and a cache file name, load the cache file and
    return the cache as a DataFrame and a set of Path objects to process.
    """
    all_jpg_file_paths = read_list_of_files(folder_to_process)
    # Take just the name, no folder into
    all_jpg_files = {f.name for f in all_jpg_file_paths}

    if os.path.exists(cache_file_path):
        # Read cache, get file names, read against new file names.
        # Remove cached file names from all_jpg_files and set file_to_process.
        with open(cache_file_path, "rb") as f:
            cached_df = pickle.load(f)

        cached_file_names = set(cached_df["image_file_name"].to_list())
        files_to_process = all_jpg_files - cached_file_names
    else:
        # No cache, so process all files.
        cached_df = None
        files_to_process = all_jpg_files

    file_paths_to_process = {Path(folder_to_process, f) for f in files_to_process}
    return cached_df, file_paths_to_process


def calc_overall_times(output_from_ocr: dict) -> list[str]:
    """
    Given the output from Mindee, calculate the overall time.
    """
    lap_time_strings = [
        "0" + re.search(r"^\d:\d{2}\.\d{3}$", v).group(0) + "000"
        for k, v in output_from_ocr.items()
        if len(k) > 2
        and re.search(r"time$", k) is not None
        and v is not None
        and re.search(r"^\d:\d{2}\.\d{3}$", v) is not None
    ]
    if len(lap_time_strings) not in [3, 7]:
        return ""
    lap_timedeltas = [
        dt.datetime.strptime(v, "%M:%S.%f") - dt.datetime(1900, 1, 1)
        for v in lap_time_strings
    ]
    overall_timedelta = sum(lap_timedeltas, dt.timedelta(seconds=0))
    overall_times = [
        "{:02}:{:02}.{:03}".format(
            overall_timedelta.seconds // 60,
            overall_timedelta.seconds % 60,
            overall_timedelta.microseconds // 1000,
        )
    ]

    if "lap_1_time_2" in output_from_ocr:
        lap_time_2_strings = [
            "0" + re.search(r"^\d:\d{2}\.\d{3}$", v).group(0) + "000"
            for k, v in output_from_ocr.items()
            if len(k) > 2
            and re.search(r"time_2$", k) is not None
            and v is not None
            and re.search(r"^\d:\d{2}\.\d{3}$", v) is not None
        ]
        if len(lap_time_2_strings) not in [3, 7]:
            return ""
        lap_2_timedeltas = [
            dt.datetime.strptime(v, "%M:%S.%f") - dt.datetime(1900, 1, 1)
            for v in lap_time_2_strings
        ]
        overall_2_timedelta = sum(lap_2_timedeltas, dt.timedelta(seconds=0))
        overall_times.append(
            [
                "{:02}:{:02}.{:03}".format(
                    overall_2_timedelta.seconds // 60,
                    overall_2_timedelta.seconds % 60,
                    overall_2_timedelta.microseconds // 1000,
                )
            ]
        )

    return overall_times


def process_list_of_files(file_paths: list[Path]) -> list[dict]:
    """
    Given a list of files (complete paths) process the files and return a list of
    dictionaries with the processed ocr text.

    Note that each dict may not have all of the desired columns. When merging, we need
    a base record with all of the columns to ensure that all are present in the
    output and that they are in the desired order.
    """
    raw_ocr_results = [extract_fields_from_image(image_path=f) for f in file_paths]
    # Purge empty dictionaries
    ocr_results = [v for v in raw_ocr_results if v != {}]

    return ocr_results


def convert_ocr_dicts_to_polars(ocr_output: list[dict]) -> pl.DataFrame:
    """
    Takes the generated ocr output (a list of dictionaries) and converts it to a polars
    DataFrame. It starts with an empty row with all of the target columns in the target
    order. The rows with data will then fall in line and there will be no missing columns.
    """
    field_names_in_order = [
        "race_end_datetime",
        "track",
        "racer",
        "character_name",
        "vehicle",
        "wheels",
        "glider",
        "overall_time",
        "lap_1_time",
        "lap_2_time",
        "lap_3_time",
        "nationality",
        "racer_2",
        "character_name_2",
        "vehicle_2",
        "wheels_2",
        "glider_2",
        "overall_time_2",
        "lap_1_time_2",
        "lap_2_time_2",
        "lap_3_time_2",
        "nationality_2",
        "image_file_name",
        "kind",
        "datetime_processed",
    ]
    default_row = [{k: "" for k in field_names_in_order}]
    # Create DataFrame with default row (which has empty strings) and given DataFrame.
    # Using the default row first ensures that all columns are present in the output
    # and that they're in the desired order.

    output_df = pl.DataFrame(default_row + ocr_output)
    # Filter to select rows with image_file_name not empty.
    output_df = output_df.filter(pl.col("image_file_name") != "")
    return output_df


def update_cache(
    cache_file_path: Path, new_df: pl.DataFrame, old_cache_df: pl.DataFrame = None
) -> None:
    """
    Given a Path for a cached pickle file of a polars DataFrame,
    a polars DataFrame of new records, and a polars DataFrame previously
    read from the cache, merge the two and save the updated DataFrame to
    the pickle file.
    """
    if old_cache_df is None:
        with open(cache_file_path, "wb") as f:
            pickle.dump(new_df, f)
        return None
    else:
        cached_and_new_df = pl.concat([old_cache_df, new_df])
        with open(cache_file_path, "wb") as f:
            pickle.dump(cached_and_new_df, f)
        return None


def set_column_autofit(worksheet):
    def as_text(value):
        if value is None:
            return ""
        return str(value)

    for column_cells in worksheet.columns:
        length = max(
            len(as_text(cell.value)) for cell in column_cells
        )  # Set length to max length of cell value.
        length = min((length + 2) * 1.2, 65)  # Cap length at 65.
        worksheet.column_dimensions[column_cells[0].column_letter].width = length
    return worksheet


def output_to_excel(
    target_excel_file: Path, new_df: pl.DataFrame, cached_df: pl.DataFrame = None
) -> None:
    """
    Given a target Excel file name and a DataFrame, output the DataFrame to the Excel
    """
    if os.path.exists(target_excel_file):
        # Append to Excel (just new)
        excel_workbook = openpyxl.load_workbook(target_excel_file)
        excel_worksheet = excel_workbook["data"]

        for row in new_df.iter_rows():
            excel_worksheet.append(list(row))

        excel_workbook.save(target_excel_file)
    else:
        # Save all to Excel.

        # Create a new Excel workbook and worksheet.
        excel_workbook = openpyxl.workbook.Workbook()
        excel_workbook.create_sheet(title="data")
        excel_workbook.remove(excel_workbook["Sheet"])
        excel_worksheet = excel_workbook["data"]

        # Write the data to the Excel worksheet
        rows = dataframe_to_rows(new_df.to_pandas(), index=False, header=True)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                excel_worksheet.cell(row=r_idx, column=c_idx, value=value)

        # Fix width of columns
        set_column_autofit(excel_worksheet)

        # Make the data into an Excel Table
        data_range = excel_worksheet.calculate_dimension()
        worksheet_table = openpyxl.worksheet.table.Table(
            displayName="data", ref=data_range
        )
        # Code to set table style doesn't seem to be working. Test by opening in Excel, not LibreOffice.
        table_style = openpyxl.worksheet.table.TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=True,
            showLastColumn=True,
            showRowStripes=True,
            showColumnStripes=True,
        )
        worksheet_table.tableStyleInfo = table_style
        excel_worksheet.add_table(worksheet_table)

        excel_workbook.save(target_excel_file)
    return None


def main(
    folder_to_process: Path,
    cache_file_path: Path,
    target_excel_path: Path,
    verbose: bool = False,
    debug: bool = False,
) -> None:
    """
    Given a folder to process, a cache file name, and a target Excel file name, process the
    files in the folder, update the cache, and update the Excel file.
    """
    if verbose:
        print("Starting main function.")

    cached_df, files_to_process = load_cache_and_list_of_paths_to_process(
        folder_to_process, cache_file_path
    )
    if len(files_to_process) == 0:
        if verbose:
            print("No files to process.")
        return None
    if verbose:
        print(f"Files to process: {files_to_process}")

    processed_ocr_output = process_list_of_files(files_to_process)
    if verbose:
        print("Processed OCR output:")
        pprint(processed_ocr_output)

    new_records_df = convert_ocr_dicts_to_polars(processed_ocr_output)
    if verbose:
        print("Number of new records:", new_records_df.shape[0])

    update_cache(cache_file_path, new_records_df, cached_df)
    if verbose:
        print("Cache updated.")

    output_to_excel(target_excel_path, new_records_df, cached_df)
    if verbose:
        print("Excel updated.")

    return None


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-f",
        "--folder_to_process",
        type=str,
        help="Path to the folder where the images are stored",
    )
    parser.add_argument(
        "-c", "--cache_file_name", type=str, help="Full path and name of the cache file"
    )
    parser.add_argument(
        "-t",
        "--target_excel_file",
        type=str,
        help="Full path and name of the target Excel file",
    )
    parser.add_argument("--verbose", action="store_true", help="Enable verbose output")
    parser.add_argument(
        "--debug", action="store_true", help="Debug mode: use pre-generated OCR results"
    )
    args = parser.parse_args()

    debug = True if args.debug else False

    main(
        folder_to_process=Path(args.folder_to_process),
        cache_file_path=Path(args.cache_file_name),
        target_excel_path=Path(args.target_excel_file),
        verbose=args.verbose,
        debug=debug,
    )

    # python ./update_mk8dx_excel_data.py -f ./data/test_images -c ./cache.pkl -t ./MK8DX_lap_times.xlsx --verbose --debug
